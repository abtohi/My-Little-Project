import pandas as pd
from openpyxl.workbook import Workbook
import os
import numpy as np
import warnings
from tqdm import tqdm

class Yuzu:
    def __init__(self,par_file,sheet):
        param          = pd.read_excel(par_file,sheet_name=sheet)
        param.columns  = list(param.iloc[2,])
        act_type       = param.iloc[0,0]
        self.period    = param.iloc[0,1]
        self.adv_share = param.iloc[0,2]
        param          = param[3:].dropna(axis=1, how='all')
        
        param.reset_index(inplace = True, drop = True)
        
        self.param      = param
        self.out_folder = 'Output'
        
        check_dir = os.path.basename(os.path.normpath(os.getcwd()))
        if(check_dir == 'Files'):
            self.file_list = os.listdir()
        else:
            os.chdir('Files')
            self.file_list = os.listdir()

        os.chdir('..')
        isExist = os.path.exists(self.out_folder)
        if isExist == False:
            os.mkdir(self.out_folder)
            
        if act_type == 'AVA SCORE':
            self.ava_score()
        elif act_type == 'AVA DISTRIBUSI':
            self.ava_distribusi()
        elif act_type == 'ADVOCACY SHARE':
            self.advocacy_share()
        elif act_type == 'PRODUCT STOCK':
            self.product_stock()
        else:
            pass
    
    def ava_score(self):
        
        for file in self.file_list:
            wb = Workbook()
            re = file.replace('.xlsx','')
            wb.save(self.out_folder+'/'+re+' - AVA SCORE.xlsx')
            writer = pd.ExcelWriter(self.out_folder+'/'+re+' - AVA SCORE.xlsx', engine='xlsxwriter')

            for i, col in self.param.iterrows():
                #Data Preprocessing
                df         = pd.read_excel('Files/'+re+'.xlsx',sheet_name=self.param.loc[i,'sheets'], skiprows=3)[3:]
                df.reset_index(inplace = True, drop = True)

                df         = df[df['Unnamed: 0'].str.len() > 0]
                df         = df.dropna(axis=1, how='all')
                df.columns = df.columns.str.replace('\n',' ').str.replace('INNER',' INNER')
                df.columns = df.columns.str.replace('  ',' ')
                len_rsa    = len(df.iloc[0,4:])
                n_rsa      = list(np.arange(4,4+len_rsa))
                rsa_ava    = df.iloc[:,[0] + n_rsa]

                rsa_ava    = rsa_ava.melt(id_vars='Unnamed: 0',var_name='kota',value_name='total')
                rsa_ava    = rsa_ava.rename(columns={'Unnamed: 0': 'provider'})
                ref        = pd.read_excel('Template/AVA SCORE.xlsx')
                ref        = ref.columns.str.strip()
                header     = pd.DataFrame(columns=ref)

                header.drop(
                    [
                        'Column 0',
                        'dashboard',
                        'Periode', 
                        'ROPERATOR AVA', 
                        'ROPERATOR new',
                        'AVA',
                        'AVA Score',
                        'AVA Score new',
                        'RSA AVA',
                        'Region AVA',
                        'Area AVA'
                    ], axis=1, inplace=True)

                header['ROPERATOR AVA'] = rsa_ava['provider']
                header['ROPERATOR new'] = rsa_ava['provider']
                header['AVA Score']     = rsa_ava['total']
                header['AVA Score new'] = rsa_ava['total']
                header['dashboard']     = 'AVA'
                header['Column 0']      = np.arange(1,len(rsa_ava)+1)
                header['Periode']       = self.period
                header['RSA AVA']       = rsa_ava['kota']
                header['Region AVA']    = df.columns[2]
                header['Area AVA']      = df.columns[3]

                if self.param.loc[i,'sheets']:
                    header['AVA'] = self.param.loc[i,'ava_type']

                header = header[list(ref)]          
                header.to_excel(writer, sheet_name=self.param.loc[i,'sheets'], index=False)

            writer.save()
            
    def ava_distribusi(self):
        for file in self.file_list:
            wb = Workbook()
            re = file.replace('.xlsx','')
            wb.save(self.out_folder+'/'+re+' - AVA DISTRIBUSI.xlsx')
            writer = pd.ExcelWriter(self.out_folder+'/'+re+' - AVA DISTRIBUSI.xlsx', engine='xlsxwriter')

            for i, col in self.param.iterrows():
                #Data Preprocessing
                df         = pd.read_excel('Files/'+re+'.xlsx',sheet_name=self.param.loc[i,'sheets'], skiprows=3)[3:]
                df.reset_index(inplace = True, drop = True)

                df         = df[df['Unnamed: 0'].str.len() > 0]
                df         = df.dropna(axis=1, how='all')
                df.columns = df.columns.str.replace('\n',' ').str.replace('INNER',' INNER')
                df.columns = df.columns.str.replace('  ',' ')
                len_rsa    = len(df.iloc[0,4:])
                n_rsa      = list(np.arange(4,4+len_rsa))
                rsa_ava    = df.iloc[:,[0] + n_rsa]

                rsa_ava    = rsa_ava.melt(id_vars='Unnamed: 0',var_name='kota',value_name='total')
                rsa_ava    = rsa_ava.rename(columns={'Unnamed: 0': 'provider'})
                ref        = pd.read_excel('Template/AVA DISTRIBUSI.xlsx')
                ref        = ref.columns.str.strip()
                header     = pd.DataFrame(columns=ref)

                header.drop(
                    [
                        'Column 0',
                        'dashboard',
                        'Periode',
                        'Kategori',
                        'ROPERATOR AVA', 
                        'AVA',
                        'AVA Score',
                        'RSA AVA',
                        'Region AVA',
                        'Area AVA'
                    ], axis=1, inplace=True)

                header['ROPERATOR AVA'] = rsa_ava['provider']
                header['AVA Score']     = rsa_ava['total']
                header['dashboard']     = 'AVA Distribusi'
                header['Column 0']      = np.arange(1,len(rsa_ava)+1)
                header['Periode']       = self.period
                header['RSA AVA']       = rsa_ava['kota']
                header['Region AVA']    = df.columns[2]
                header['Area AVA']      = df.columns[3]

                if self.param.loc[i,'sheets']:
                    header['AVA']      = self.param.loc[i,'ava_type']
                    header['Kategori'] = self.param.loc[i,'category']

                header = header[list(ref)]          
                header.to_excel(writer, sheet_name=self.param.loc[i,'sheets'], index=False)

            writer.save()
            
    def advocacy_share(self):
        for file in self.file_list:
            wb = Workbook()
            re = file.replace('.xlsx','')
            wb.save(self.out_folder+'/'+re+' - ADVOCACY SHARE.xlsx')
            writer = pd.ExcelWriter(self.out_folder+'/'+re+' - ADVOCACY SHARE.xlsx', engine='xlsxwriter')
            final  = pd.DataFrame()

            for i, col in tqdm(self.param.iterrows()):
                #Data Preprocessing
                df         = pd.read_excel('Files/'+re+'.xlsx',sheet_name=self.param.loc[i,'sheets'], skiprows=3)[3:]
                df.reset_index(inplace = True, drop = True)

                df         = df[df['Unnamed: 0'].str.len() > 0]
                df         = df.dropna(axis=1, how='all')
                df.columns = df.columns.str.replace('\n',' ').str.replace('INNER',' INNER')
                df.columns = df.columns.str.replace('  ',' ')
                len_rsa    = len(df.iloc[0,4:])
                n_rsa      = list(np.arange(4,4+len_rsa))
                rsa_ava    = df.iloc[:,[0] + n_rsa]

                rsa_ava    = rsa_ava.melt(id_vars='Unnamed: 0',var_name='kota',value_name='total')
                rsa_ava    = rsa_ava.rename(columns={'Unnamed: 0': 'provider'})
                ref        = pd.read_excel('Template/ADVOCACY SHARE.xlsx')
                ref        = ref.columns.str.strip()
                header     = pd.DataFrame(columns=ref)

                header.drop(
                    [
                        'Column 0',
                        'dashboard',
                        'RSA',
                        'Periode',
                        'Kategori',
                        'Unnamed  0',
                        'Variable AS',
                        'AS Score',
                        'Advocacy Share',
                        'Region AS',
                        'Area AS',
                        'file'   
                    ], axis=1, inplace=True)
                
                header['Column 0']       = np.arange(1,len(rsa_ava)+1)
                header['dashboard']      = 'Advocacy Share'
                header['RSA']            = rsa_ava['kota']
                header['Periode']        = self.period
                header['Unnamed  0']     = np.arange(1,len(rsa_ava)+1)
                header['Variable AS']    = rsa_ava['provider']
                header['AS Score']       = rsa_ava['total']             
                header['Region AS']      = df.columns[2]
                header['Area AS']        = df.columns[3]
                header['file']           = re+'.xlsx'

                if self.param.loc[i,'sheets']:
                    header['Kategori'] = self.param.loc[i,'category']
                    header['Advocacy Share'] = self.param.loc[i,'advocacy_share']

                header = header[list(ref)]
                final = final.append(header)        
            
            final.reset_index(drop=True)
            final['Column 0'] = np.arange(1,len(final)+1)
            final['Unnamed  0'] = np.arange(1,len(final)+1)
            final.to_excel(writer, sheet_name='Output', index=False)
            writer.save()
      
    def product_stock(self):
        for file in self.file_list:
            wb = Workbook()
            re = file.replace('.xlsx','')
            wb.save(self.out_folder+'/'+re+' - PRODUCT STOCK.xlsx')
            writer = pd.ExcelWriter(self.out_folder+'/'+re+' - PRODUCT STOCK.xlsx', engine='xlsxwriter')

            for i, col in self.param.iterrows():
                #Data Preprocessing
                df         = pd.read_excel('Files/'+re+'.xlsx',sheet_name=self.param.loc[i,'sheets'], skiprows=3)[3:]
                df.reset_index(inplace = True, drop = True)

                df         = df[df['Unnamed: 0'].str.len() > 0]
                df         = df.dropna(axis=1, how='all')
                df.columns = df.columns.str.replace('\n',' ').str.replace('INNER',' INNER')
                df.columns = df.columns.str.replace('  ',' ')
                len_rsa    = len(df.iloc[0,4:])
                n_rsa      = list(np.arange(4,4+len_rsa))
                rsa_ava    = df.iloc[:,[0] + n_rsa]

                rsa_ava    = rsa_ava.melt(id_vars='Unnamed: 0',var_name='kota',value_name='total')
                rsa_ava    = rsa_ava.rename(columns={'Unnamed: 0': 'provider'})
                ref        = pd.read_excel('Template/PRODUCT STOCK.xlsx')
                ref        = ref.columns.str.strip()
                header     = pd.DataFrame(columns=ref)

                header.drop(
                    [
                        'Column 0',
                        'dashboard',
                        'Periode',
                        'SA',
                        'ROPERATOR PS',
                        'RSA PS',
                        'PS Score',
                        'Kategori PS',
                        'Area PS',
                        'Region PS',
                        'Unnamed  0'
                    ], axis=1, inplace=True)
                
                header['Column 0']       = np.arange(1,len(rsa_ava)+1)
                header['dashboard']      = 'Product Stock'
                header['Periode']        = self.period
                header['SA']             = rsa_ava['kota'].str.upper()
                header['ROPERATOR PS']   = rsa_ava['provider']
                header['RSA PS']         = rsa_ava['kota']
                header['PS Score']       = rsa_ava['total']
                header['Area PS']        = df.columns[3]
                header['Region PS']      = df.columns[2]
                header['Unnamed  0']     = np.arange(1,len(rsa_ava)+1)             

                if self.param.loc[i,'sheets']:
                    header['Kategori PS'] = self.param.loc[i,'category']

                header = header[list(ref)]          
                header.to_excel(writer, sheet_name=self.param.loc[i,'sheets'], index=False)

            writer.save()
            
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
run_this = Yuzu('parameters.xlsx','PARAM')
